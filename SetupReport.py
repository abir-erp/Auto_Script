# importing pandas as pd

from operator import index
import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill

# read an excel file and convert

inp="C:/Setup_code/report_setup_times.xlsx"
output="C:/Setup_code/temp.xlsx"


xlLocMain=inp
xlLoc=output
# into a dataframe object
df = pd.DataFrame(pd.read_excel(xlLocMain))


df.to_excel(xlLoc,index=False)

#Sorting DONE

df = pd.DataFrame(pd.read_excel(xlLoc))
t=[]

for i in range(len(df.index)):
    t.append([df.iloc[i,2],df.iloc[i,3],df.iloc[i,4]])

temp = []
for x in t:
    if x not in temp:
        temp.append(x)
t = temp

#df3 is temporary
df3 = df.drop(df.columns[[0, 1, 5,6,7,8,9,10,11,12,13,14,15,16,17]], axis=1)
#print(df3)
df3=df3.values.tolist()
temp_l=[]
for c in df3:
    if c in t:
        temp_l.append( (t.index(c))+1 )
#print(temp,temp_l)
#df['ID'] = temp_l

if 'ID' not in df:
    df.insert(loc=0, column='ID', value=temp_l)
df.to_excel(xlLoc,index=False)

df = pd.DataFrame(pd.read_excel(xlLoc))
df['Date'] = pd.to_datetime(df['Date']).dt.date
df=df.sort_values(by=['ID','Date'],ascending=True)
if 'Day Count' not in df:
    df.insert(2,"Day Count","")
if 'Time Spent' not in df: 
    df.insert(4,"Time Spent","")
#df=df.groupby(['ID','Part Number']).sum('Employee')
df.to_excel(xlLoc,index=False)
#print(t,len(t),df3,len(df3))
col_to_move = 'Part Number'
new_position = 1
cols = list(df.columns)
cols.insert(new_position, cols.pop(cols.index(col_to_move)))
df = df[cols]
print(df)
#
# 
# #
listOfDates = list(dict.fromkeys(df['Date'].tolist()))#***************
listOfDates.sort()
listOfDates.reverse()
listOfID = list(dict.fromkeys(df['ID'].tolist()))#***************

df=df.sort_values(by=['Date','Part Number','Op Number','SO Number'],ascending=False)
#print(listOfDates,'\nhahaha\n',listOfID)

df = pd.DataFrame(pd.read_excel(xlLoc))
df_list=df.values.tolist()
#df=df.loc[df.n == "d", ['ID',	'Employee',	'Day Count',	'Date',	'Time Spent',	'Part Number',	'Op Number',	'SO Number',	'Set-Up Time',	'Burden Time',	'Operation Complete',	'SetupComplete']].values.flatten().tolist()
#print(df_list)

d={}
#print(d)
for rows in df_list:
    if rows[0] in d:
        d[rows[0]]=d[rows[0]]+[rows[3]]
    else:
        d[rows[0]]=[rows[3]]
"""    for id in listOfID:
        if id in rows:
            for dates in listOfDates:
                if dates in rows:
                    if id in d:
                        d[id]=d[id]+[dates]
                    else:
                        d[id]=[dates]"""
#print(d)
d2={}

def Merge(dict1, dict2):
    res={**dict1,**dict2}
    return res

for dates in listOfDates:
    for id in d.keys():
        if any(ts.date() == pd.Timestamp(dates).date() for ts in d[id]):
            count = sum(1 for ts in d[id] if ts.date() == pd.Timestamp(dates).date())
            if dates in d2:
                d2[dates]=Merge(d2[dates],{id:count})
            else:
                d2[dates]={id:count}
for kk,vv in d2.items():
    d2[kk]={k: v for k, v in sorted(vv.items(), key=lambda item: item[1],reverse=True)}
#print(d2)
idlistFinal=[]
ll=[]
countor=0
for valueD in d2.values():
    ll.append(list(valueD.keys()))
    

#print(idlistFinal)    
#print(ll)
ll=[item for sublist in ll for item in sublist]
sortedOrderlist = [] 
[sortedOrderlist.append(x) for x in ll if x not in sortedOrderlist] 
#sortedOrderlist=ll#List to create new excel in a proper sorted order

#print(sortedOrderlist)

#Writing in excel in a sorted order
rslt_df=pd.DataFrame()
for ids in sortedOrderlist: 
    rslt_df=pd.concat([rslt_df,df[df['ID'] == ids]])

arr = rslt_df["ID"].to_list()
result = [i for n, i in enumerate(arr) if i not in arr[:n]] 
newl=[]
for i in arr:
    newl.append(result.index(i)+1)

rslt_df['ID']=newl
rslt_df=rslt_df.sort_values(by=['ID','Date'],ascending=[True,True])
rslt_df['Date'] = pd.to_datetime(df['Date']).dt.date
rslt_df.to_excel(xlLoc,index=False)
#
# 
# #

from openpyxl import load_workbook
clr_background = PatternFill(start_color='00F0F8FF', end_color='00F0F8FF', fill_type="solid")
file = xlLoc
# Load workbook
wb = load_workbook(filename=file)
ws = wb['Sheet1']
mylist = []
for cell in ws['A']:
    mylist.append(cell.value)
print(mylist)
counter=0
for i in ws:
    if counter==0:
        print('')
        counter+=1
    else:
        t=(i[0].value)
        if (t%2)!=0:
            for j in range(21):
                i[j].fill=clr_background
        
    

mergecount=0
startcell=2
for row in range(1, len(mylist)):
    #print(row, mylist[row-1], mylist[row])
    if mylist[row-1] == mylist[row]:
        mergecount += 1
    else:
        #print(row, mylist[row-1], mylist[row], startcell, mergecount)
        for kk in range(startcell,startcell+mergecount+1):
            if kk==startcell:
                ws[f'c{kk}']=1
                ws[f'e{kk}']=f'=SUM(I{startcell}:I{mergecount+startcell})+SUM(J{startcell}:J{mergecount+startcell})'
            else:
                ws[f'c{kk}']=f'=IF(D{kk}=D{kk-1},c{kk-1},(C{kk-1}+1))'
                #print(kk,'**')
                ws[f'e{kk}']=f'=SUM(I{startcell}:I{mergecount+startcell})+SUM(J{startcell}:J{mergecount+startcell})'
        if mergecount > 0:
            #ws[kk][0].fill = clr_background
            #print(startcell,startcell+mergecount)
            
            ws.merge_cells(start_row=startcell, start_column=1, end_row=startcell+mergecount, end_column=1)
        mergecount = 0
        startcell = row+1
if mergecount > 0:
    ws.merge_cells(start_row=startcell, start_column=1, end_row=startcell+mergecount, end_column=1)
wb.save(file)




#formatting xl

mylist = []
for cell in ws['E']:
    mylist.append(cell.value)
mergecount=0
startcell=1
for row in range(1, len(mylist)):
    #print(row, mylist[row-1], mylist[row])
    if mylist[row-1] == mylist[row]:
        mergecount += 1
        #print(mergecount)
    else:
        #print(row, mylist[row-1], mylist[row], startcell, mergecount)
        if mergecount > 0:
            ws.merge_cells(start_row=startcell, start_column=5, end_row=startcell+mergecount, end_column=5)
        mergecount = 0
        startcell = row+1
    
if mergecount > 0:
    ws.merge_cells(start_row=startcell, start_column=5, end_row=startcell+mergecount, end_column=5)



mylist = []
for cell in ws['F']:
    mylist.append(cell.value)
mergecount=0
startcell=1
for row in range(1, len(mylist)):
    #print(row, mylist[row-1], mylist[row])
    if mylist[row-1] == mylist[row]:
        mergecount += 1
        #print(mergecount)
    else:
        #print(row, mylist[row-1], mylist[row], startcell, mergecount)
        if mergecount > 0:
            
            ws.merge_cells(start_row=startcell, start_column=6, end_row=startcell+mergecount, end_column=6)
        mergecount = 0
        startcell = row+1
    
if mergecount > 0:
    ws.merge_cells(start_row=startcell, start_column=6, end_row=startcell+mergecount, end_column=6)


from itertools import chain
# Using `ws` as the Worksheet

for cell in chain.from_iterable(ws.iter_cols()):
    if cell.value:
        ws.column_dimensions[cell.column_letter].width = max(
            ws.column_dimensions[cell.column_letter].width,
            len(f"{cell.value}"),
        )


"""for i in ws:
    t=i[0].value
    if t==None:
        print(t,type(t))
    elif (t % 2) == 0:
        print(t,type(t))
    #i.fill = clr_background"""

wb.save(file)


#***************************------------------------------------#